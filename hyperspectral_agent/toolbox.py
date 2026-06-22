"""
高光谱专业工具箱 (Hyperspectral Toolbox)
==========================================
基于 rasterio / numpy / scipy 的高光谱图像分层处理工具集。

所有工具面向 [Bands, Height, Width] 三维栅格数组设计，
重点支持土壤剖面高光谱图像的分层清洗与光谱分析。

工具列表：
  - load_tiff         : 加载 TIFF + 分层元数据
  - get_layer_info    : 查询当前数据的分层信息
  - extract_layer     : 提取指定土层的像素矩阵
  - extract_spectrum  : 提取指定像元的光谱曲线
  - clean_layer_noise : 分层噪声过滤
  - msc_transform     : 多元散射校正 (MSC)
  - snv_transform     : 标准正态变量变换 (SNV)
  - sg_smooth         : Savitzky-Golay 光谱平滑
  - render_heatmap    : 绘制空间丰度热力图
  - render_spectrum   : 绘制光谱响应曲线
  - compute_layer_stats: 计算分层统计特征
"""

import json
import os
from pathlib import Path
from typing import Any, Optional

import numpy as np
import rasterio
from scipy.signal import savgol_filter
from scipy.ndimage import median_filter
import os
os.environ.setdefault("MPLCONFIGDIR", os.path.join(os.path.dirname(__file__), "..", ".mpl_cache"))
import matplotlib
matplotlib.use("Agg")  # 非交互式后端
import matplotlib.pyplot as plt

# 配置中文字体
_font_candidates = ['Microsoft YaHei', 'SimHei', 'Noto Sans SC', 'KaiTi', 'SimSun']
for _fn in _font_candidates:
    try:
        plt.rcParams['font.sans-serif'] = [_fn] + plt.rcParams['font.sans-serif']
        plt.rcParams['axes.unicode_minus'] = False
        break
    except Exception:
        continue


# ═══════════════════════════════════════════════════════════════
# 数据容器
# ═══════════════════════════════════════════════════════════════

class HyperSpectralData:
    """高光谱数据容器：保存三维数组和分层元数据"""

    def __init__(self):
        self.data: Optional[np.ndarray] = None        # [Bands, Height, Width] float32
        self.band_count: int = 0
        self.height: int = 0
        self.width: int = 0
        self.profile_id: str = ""
        self.tif_path: str = ""
        self.layers: list = []  # 分层列表，每项 {"layer_id", "name", "code", "top_cm", "bottom_cm", ...}
        self.profile_depth_cm: float = 0.0  # 剖面总深度 (cm)
        self.pixel_per_cm: float = 0.0       # 每厘米像素数 (高度方向)

    def layer_count(self) -> int:
        return len(self.layers)

    def depth_to_row(self, depth_cm: float) -> int:
        """将深度 (cm) 映射到图像行索引"""
        if self.pixel_per_cm <= 0:
            return 0
        row = int(depth_cm * self.pixel_per_cm)
        return max(0, min(row, self.height - 1))

    def get_layer_pixel_range(self, layer: dict) -> tuple:
        """获取图层对应的像素行范围 (row_start, row_end)"""
        top = self.depth_to_row(layer.get("top_cm", 0))
        bottom = self.depth_to_row(layer.get("bottom_cm", 0))
        return (top, bottom)


# ═══════════════════════════════════════════════════════════════
# 全局数据实例
# ═══════════════════════════════════════════════════════════════

_hs_data = HyperSpectralData()


def get_data() -> HyperSpectralData:
    """获取当前高光谱数据"""
    return _hs_data


# ═══════════════════════════════════════════════════════════════
# 工具 1: 加载 TIFF + 分层元数据
# ═══════════════════════════════════════════════════════════════

def load_tiff(tif_path: str, meta_path: str = "") -> str:
    """
    加载高光谱 TIFF 文件及分层元数据。

    参数:
        tif_path  : .tif / .tiff 文件路径
        meta_path : 分层配置文件路径 (.xlsx / .json / .txt)
                    - .xlsx: Excel 格式，列必须包含 编号/层次/发生层名称/发生层顶部/发生层底部
                    - .json: {"profile_id": "...", "layers": [...]}
    """
    try:
        # ── 读取 TIFF ──
        with rasterio.open(tif_path) as src:
            _hs_data.data = src.read()  # [Bands, Height, Width]
            _hs_data.band_count = src.count
            _hs_data.height = src.height
            _hs_data.width = src.width
            _hs_data.tif_path = tif_path

        # ── 读取分层元数据 ──
        if meta_path:
            ext = os.path.splitext(meta_path)[1].lower()
            if ext in (".xlsx", ".xls"):
                _load_layers_from_xlsx(meta_path)
            elif ext == ".json":
                _load_layers_from_json(meta_path)
            elif ext == ".txt":
                _load_layers_from_txt(meta_path)
            else:
                return f"不支持的分层文件格式: {ext}，支持 .xlsx / .json / .txt"

        # ── 用 profile_id 过滤该 TIF 对应的分层 ──
        profile_id = _hs_data.profile_id
        if profile_id and meta_path:
            _hs_data.layers = [ly for ly in _hs_data.layers if ly.get("profile_id") == profile_id]
            if not _hs_data.layers:
                return (f"警告: 在元数据中未找到编号 '{profile_id}' 的分层信息。"
                        f"请确认 TIF 文件夹名与 Excel 的'编号'列一致。")

        # ── 计算像素/厘米比例 ──
        if _hs_data.layers:
            # 取最底层底部深度作为剖面总深度
            max_bottom = max(ly.get("bottom_cm", 0) for ly in _hs_data.layers)
            _hs_data.profile_depth_cm = max_bottom
            if max_bottom > 0:
                _hs_data.pixel_per_cm = _hs_data.height / max_bottom

        layers_desc = "\n".join(
            f"  - 层次{ly.get('layer_id','?')}: {ly.get('name','?')} "
            f"({ly.get('code','?')}) [{ly.get('top_cm',0)}-{ly.get('bottom_cm',0)}cm]"
            for ly in _hs_data.layers
        )

        return (
            f"TIFF 加载成功。\n"
            f"  波段数: {_hs_data.band_count}, "
            f"空间尺寸: {_hs_data.width}×{_hs_data.height} 像素\n"
            f"  数据范围: [{_hs_data.data.min():.4f}, {_hs_data.data.max():.4f}]\n"
            f"  分层数: {_hs_data.layer_count()}\n"
            f"{layers_desc}"
        )

    except Exception as e:
        return f"加载失败: {str(e)}"


def _infer_profile_id(tif_path: str) -> str:
    """TIF parent folder name as profile_id"""
    parent = os.path.basename(os.path.dirname(os.path.abspath(tif_path)))
    return parent


def find_tif_in_folder(folder_path: str) -> Optional[str]:
    """Auto-discover .tif/.tiff file in a folder"""
    folder = Path(folder_path)
    if not folder.is_dir():
        return None
    for ext in ('.tif', '.tiff', '.TIF', '.TIFF'):
        candidates = list(folder.glob(f'*{ext}'))
        if candidates:
            return str(candidates[0])
    return None


def set_meta_file(meta_path: str) -> str:
    """Set global metadata file path and persist to config"""
    _hs_data._meta_path = meta_path
    # Also persist to config
    try:
        from .config import AppConfig
        cfg = AppConfig.from_file()
        cfg.meta_file_path = meta_path
        cfg.save()
    except Exception:
        pass
    return f'Metadata file set: {meta_path}'


def load_from_folder(folder_path: str) -> str:
    """
    Select a profile folder. Uses folder name as profile ID,
    auto-discovers the .tif inside, and matches layers from pre-set metadata.
    """
    folder_path = os.path.abspath(folder_path)
    folder_name = os.path.basename(folder_path)

    tif_path = find_tif_in_folder(folder_path)
    if not tif_path:
        return f'Error: No .tif file found in {folder_path}'

    meta_path = getattr(_hs_data, '_meta_path', '')
    _hs_data.profile_id = folder_name

    return load_tiff(tif_path, meta_path)


def _load_layers_from_xlsx(xlsx_path: str) -> None:
    """从 Excel 文件加载分层信息"""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
    current_profile = ""
    layers = []
    all_layers = []

    for row in rows:
        # 列: 序号, 编号, 层次, 发生层, 发生层名称, 发生层顶部, 发生层底部, ...
        seq = str(row[0]).strip() if row[0] is not None else ""
        profile_id = str(row[1]).strip() if row[1] is not None else ""
        layer_id = str(row[2]).strip() if row[2] is not None else ""
        code = str(row[3]).strip() if row[3] is not None else ""
        name = str(row[4]).strip() if row[4] is not None else ""
        top_cm = float(row[5]) if row[5] is not None else 0.0
        bottom_cm = float(row[6]) if row[6] is not None else 0.0

        # 合并裂开的多列（如: 裂缝/结核/石头/树根 在列 9-12）
        annotations = []
        for col_idx in range(9, min(len(row), 15)):
            val = row[col_idx]
            if val is not None and str(val).strip():
                annotations.append(str(val).strip())

        if seq:  # 新的剖面
            if current_profile and layers:
                all_layers.append((current_profile, layers))
            current_profile = profile_id
            layers = []

        layers.append({
            "profile_id": profile_id if profile_id else current_profile,
            "layer_id": layer_id,
            "code": code,
            "name": name,
            "top_cm": top_cm,
            "bottom_cm": bottom_cm,
            "annotations": annotations,
        })

    if current_profile and layers:
        all_layers.append((current_profile, layers))

    wb.close()

    # 推断当前 TIF 对应的 profile_id
    _hs_data.profile_id = _infer_profile_id(_hs_data.tif_path)

    # 存储所有分层（后续按 profile_id 过滤）
    _hs_data._all_layers = all_layers
    _hs_data.layers = []
    for pid, lys in all_layers:
        _hs_data.layers.extend(lys)


def _load_layers_from_json(json_path: str) -> None:
    """从 JSON 文件加载分层信息"""
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    _hs_data.profile_id = data.get("profile_id", _infer_profile_id(_hs_data.tif_path))
    _hs_data.layers = data.get("layers", [])

    # 补充默认值
    for ly in _hs_data.layers:
        ly.setdefault("profile_id", _hs_data.profile_id)
        ly.setdefault("layer_id", "")
        ly.setdefault("code", "")
        ly.setdefault("name", "")
        ly.setdefault("top_cm", 0.0)
        ly.setdefault("bottom_cm", 0.0)
        ly.setdefault("annotations", [])


def _load_layers_from_txt(txt_path: str) -> None:
    """从简单 TXT 文件加载分层（每行: 层次编号,名称,顶部cm,底部cm）"""
    layers = []
    profile_id = _infer_profile_id(_hs_data.tif_path)
    _hs_data.profile_id = profile_id

    with open(txt_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = [p.strip() for p in line.split(",")]
            if len(parts) >= 4:
                layers.append({
                    "profile_id": profile_id,
                    "layer_id": parts[0],
                    "name": parts[1],
                    "top_cm": float(parts[2]),
                    "bottom_cm": float(parts[3]),
                    "code": "",
                    "annotations": [],
                })
    _hs_data.layers = layers


# ═══════════════════════════════════════════════════════════════
# 工具 2: 获取分层信息
# ═══════════════════════════════════════════════════════════════

def get_layer_info() -> str:
    """返回当前加载数据的分层信息摘要"""
    if not _hs_data.layers:
        return "尚未加载分层数据，请先调用 load_tiff。"
    lines = [f"剖面编号: {_hs_data.profile_id}"]
    lines.append(f"剖面深度: {_hs_data.profile_depth_cm} cm, 像素比例: {_hs_data.pixel_per_cm:.2f} px/cm")
    lines.append(f"共 {_hs_data.layer_count()} 层:")
    for ly in _hs_data.layers:
        extra = f" 标注: {ly['annotations']}" if ly.get("annotations") else ""
        lines.append(
            f"  层次 {ly['layer_id']}: {ly['name']} ({ly['code']}) "
            f"深度 {ly['top_cm']}-{ly['bottom_cm']} cm{extra}"
        )
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════
# 工具 3: 提取指定土层的像素矩阵
# ═══════════════════════════════════════════════════════════════

def extract_layer(layer_id: str) -> str:
    """
    提取指定土层的二维像素矩阵。

    参数:
        layer_id: 土层编号 (字符串，如 "1", "2", "3")

    返回:
        包含该层所有波段数据的统计摘要
    """
    if _hs_data.data is None:
        return "错误: 尚未加载 TIFF 数据。"

    # 查找图层
    target = None
    for ly in _hs_data.layers:
        if str(ly.get("layer_id", "")) == str(layer_id):
            target = ly
            break

    if target is None:
        return f"错误: 未找到土层 '{layer_id}'。可用土层: {[ly['layer_id'] for ly in _hs_data.layers]}"

    top, bottom = _hs_data.get_layer_pixel_range(target)
    layer_data = _hs_data.data[:, top:bottom, :]  # [Bands, LayerHeight, Width]

    # 统计信息
    mean_vals = np.mean(layer_data, axis=(1, 2))  # 每个波段的均值
    std_vals = np.std(layer_data, axis=(1, 2))

    report = (
        f"土层 '{layer_id}' ({target['name']}, {target['code']}) 提取成功。\n"
        f"  深度范围: {target['top_cm']}-{target['bottom_cm']} cm\n"
        f"  像素范围: 行 {top}-{bottom} ({bottom - top} 行)\n"
        f"  数据形状: {layer_data.shape}\n"
        f"  全波段均值范围: [{mean_vals.min():.4f}, {mean_vals.max():.4f}]\n"
        f"  全波段均值: {mean_vals.mean():.4f} ± {mean_vals.std():.4f}"
    )
    return report


# ═══════════════════════════════════════════════════════════════
# 工具 4: 提取指定像元的光谱曲线
# ═══════════════════════════════════════════════════════════════

def extract_spectrum(row: int, col: int) -> str:
    """
    提取指定像元 (row, col) 的全波段光谱反射率曲线。

    返回:
        光谱统计信息 + 峰值波段
    """
    if _hs_data.data is None:
        return "错误: 尚未加载 TIFF 数据。"

    if row < 0 or row >= _hs_data.height or col < 0 or col >= _hs_data.width:
        return f"错误: 像元坐标 ({row},{col}) 超出图像范围 ({_hs_data.height},{_hs_data.width})。"

    spectrum = _hs_data.data[:, row, col]  # [Bands]

    # 找峰值
    peak_band = int(np.argmax(spectrum))
    valley_band = int(np.argmin(spectrum))

    depth_cm = row / _hs_data.pixel_per_cm if _hs_data.pixel_per_cm > 0 else 0.0

    return (
        f"像元 ({row},{col}) 光谱曲线:\n"
        f"  对应深度: {depth_cm:.1f} cm\n"
        f"  反射率范围: [{spectrum.min():.4f}, {spectrum.max():.4f}]\n"
        f"  均值: {spectrum.mean():.4f}, 中位数: {np.median(spectrum):.4f}\n"
        f"  峰值波段: {peak_band + 1} (值: {spectrum[peak_band]:.4f})\n"
        f"  谷值波段: {valley_band + 1} (值: {spectrum[valley_band]:.4f})"
    )


# ═══════════════════════════════════════════════════════════════
# 工具 5: 分层噪声清洗
# ═══════════════════════════════════════════════════════════════

def clean_layer_noise(layer_id: str, strategy: str = "median") -> str:
    """
    对指定土层进行空间噪声过滤。

    参数:
        layer_id : 土层编号
        strategy : 过滤策略
                   "median" - 中值滤波 (去除椒盐噪声 / 死像素)
                   "mean"   - 均值插值 (填充异常值)
                   "none"   - 不处理，仅检测异常像元

    返回:
        清洗前后统计对比
    """
    if _hs_data.data is None:
        return "错误: 尚未加载 TIFF 数据。"

    target = None
    for ly in _hs_data.layers:
        if str(ly.get("layer_id", "")) == str(layer_id):
            target = ly
            break

    if target is None:
        return f"错误: 未找到土层 '{layer_id}'。"

    top, bottom = _hs_data.get_layer_pixel_range(target)
    layer_data = _hs_data.data[:, top:bottom, :].copy()

    before_mean = layer_data.mean()
    before_std = layer_data.std()

    if strategy == "median":
        # 对每个波段单独做 3×3 中值滤波
        for b in range(layer_data.shape[0]):
            layer_data[b] = median_filter(layer_data[b], size=3)

    elif strategy == "mean":
        # 用全局均值填充异常值 (超过 3σ)
        for b in range(layer_data.shape[0]):
            band = layer_data[b]
            mu, sigma = band.mean(), band.std()
            mask = np.abs(band - mu) > 3 * sigma
            band[mask] = mu

    elif strategy == "none":
        # 仅检测
        outlier_count = 0
        for b in range(layer_data.shape[0]):
            band = layer_data[b]
            mu, sigma = band.mean(), band.std()
            outlier_count += np.sum(np.abs(band - mu) > 3 * sigma)

        return (
            f"土层 '{layer_id}' 异常检测完成。\n"
            f"  异常像元数 (>3σ): {outlier_count}\n"
            f"  总像元数: {layer_data.size}\n"
            f"  异常比例: {outlier_count / layer_data.size * 100:.2f}%\n"
            f"  波段均值: {before_mean:.4f} ± {before_std:.4f}"
        )

    else:
        return f"错误: 不支持的清洗策略 '{strategy}'。可用: median, mean, none"

    # 写回数据
    _hs_data.data[:, top:bottom, :] = layer_data

    after_mean = layer_data.mean()
    after_std = layer_data.std()

    return (
        f"土层 '{layer_id}' ({target['name']}) 清洗完成 (策略: {strategy})。\n"
        f"  清洗前均值/标准差: {before_mean:.4f} / {before_std:.4f}\n"
        f"  清洗后均值/标准差: {after_mean:.4f} / {after_std:.4f}"
    )


# ═══════════════════════════════════════════════════════════════
# 工具 6: 多元散射校正 (MSC)
# ═══════════════════════════════════════════════════════════════

def msc_transform(layer_ids: str = "") -> str:
    """
    对指定土层集合进行多元散射校正 (MSC)。

    原理: 以平均光谱为参考，对每个像元的光谱做线性回归消除散射效应。

    参数:
        layer_ids : 逗号分隔的土层编号（如 "1,2,3"），
                    留空则对所有土层处理
    """
    if _hs_data.data is None:
        return "错误: 尚未加载 TIFF 数据。"

    # 确定处理范围
    if layer_ids:
        ids = [s.strip() for s in layer_ids.split(",")]
        masks = []
        for lid in ids:
            for ly in _hs_data.layers:
                if str(ly.get("layer_id", "")) == lid:
                    top, bottom = _hs_data.get_layer_pixel_range(ly)
                    mask = np.zeros(_hs_data.height, dtype=bool)
                    mask[top:bottom] = True
                    masks.append(mask)
                    break
        if not masks:
            return f"错误: 未找到指定土层 {ids}。"
        row_mask = np.any(masks, axis=0)
    else:
        row_mask = np.ones(_hs_data.height, dtype=bool)

    # 提取要处理的数据
    subset = _hs_data.data[:, row_mask, :]  # [Bands, SelectedRows, Width]
    bands, h, w = subset.shape
    spectra = subset.reshape(bands, -1).T  # [N_pixels, Bands]

    # 计算平均光谱作为参考
    mean_spectrum = spectra.mean(axis=0)

    # 对每个像元做 MSC
    corrected = np.zeros_like(spectra)
    for i in range(spectra.shape[0]):
        # 线性回归: spectrum_i = a + b * mean_spectrum
        b = np.cov(spectra[i], mean_spectrum)[0, 1] / np.var(mean_spectrum)
        a = spectra[i].mean() - b * mean_spectrum.mean()
        corrected[i] = (spectra[i] - a) / b

    # 写回
    corrected_3d = corrected.T.reshape(bands, h, w)
    _hs_data.data[:, row_mask, :] = corrected_3d

    layer_desc = f"土层 {layer_ids}" if layer_ids else "所有土层"
    return (
        f"MSC 变换完成 ({layer_desc})。\n"
        f"  处理像元数: {spectra.shape[0]}\n"
        f"  校正后均值: {corrected.mean():.4f}, 标准差: {corrected.std():.4f}"
    )


# ═══════════════════════════════════════════════════════════════
# 工具 7: 标准正态变量变换 (SNV)
# ═══════════════════════════════════════════════════════════════

def snv_transform(layer_ids: str = "") -> str:
    """
    对指定土层集合进行标准正态变量变换 (SNV)。

    原理: 对每个像元的光谱单独做标准化，消除散射和光程变化影响。

    参数:
        layer_ids : 逗号分隔的土层编号，留空则对所有土层处理
    """
    if _hs_data.data is None:
        return "错误: 尚未加载 TIFF 数据。"

    if layer_ids:
        ids = [s.strip() for s in layer_ids.split(",")]
        masks = []
        for lid in ids:
            for ly in _hs_data.layers:
                if str(ly.get("layer_id", "")) == lid:
                    top, bottom = _hs_data.get_layer_pixel_range(ly)
                    mask = np.zeros(_hs_data.height, dtype=bool)
                    mask[top:bottom] = True
                    masks.append(mask)
                    break
        if not masks:
            return f"错误: 未找到指定土层 {ids}。"
        row_mask = np.any(masks, axis=0)
    else:
        row_mask = np.ones(_hs_data.height, dtype=bool)

    subset = _hs_data.data[:, row_mask, :]
    bands, h, w = subset.shape
    spectra = subset.reshape(bands, -1).T  # [N_pixels, Bands]

    # SNV: (x - mean) / std (per pixel)
    means = spectra.mean(axis=1, keepdims=True)
    stds = spectra.std(axis=1, keepdims=True)
    stds[stds == 0] = 1.0  # 防止除零
    corrected = (spectra - means) / stds

    corrected_3d = corrected.T.reshape(bands, h, w)
    _hs_data.data[:, row_mask, :] = corrected_3d

    layer_desc = f"土层 {layer_ids}" if layer_ids else "所有土层"
    return (
        f"SNV 变换完成 ({layer_desc})。\n"
        f"  处理像元数: {spectra.shape[0]}\n"
        f"  校正后均值: {corrected.mean():.4f}, 标准差: {corrected.std():.4f}"
    )


# ═══════════════════════════════════════════════════════════════
# 工具 8: Savitzky-Golay 光谱平滑
# ═══════════════════════════════════════════════════════════════

def sg_smooth(window_length: int = 7, polyorder: int = 2, layer_ids: str = "") -> str:
    """
    沿波段轴对每个像元的光谱曲线进行 Savitzky-Golay 平滑。

    参数:
        window_length : 滑动窗口大小 (奇数, ≥ polyorder+1)
        polyorder     : 多项式阶数
        layer_ids     : 逗号分隔的土层编号，留空则处理所有土层
    """
    if _hs_data.data is None:
        return "错误: 尚未加载 TIFF 数据。"

    # 参数校验
    if window_length % 2 == 0:
        window_length += 1
    if window_length <= polyorder:
        return f"错误: window_length ({window_length}) 必须大于 polyorder ({polyorder})。"

    if window_length > _hs_data.band_count:
        window_length = _hs_data.band_count
        if window_length % 2 == 0:
            window_length -= 1

    # 确定处理范围
    if layer_ids:
        ids = [s.strip() for s in layer_ids.split(",")]
        masks = []
        for lid in ids:
            for ly in _hs_data.layers:
                if str(ly.get("layer_id", "")) == lid:
                    top, bottom = _hs_data.get_layer_pixel_range(ly)
                    mask = np.zeros(_hs_data.height, dtype=bool)
                    mask[top:bottom] = True
                    masks.append(mask)
                    break
        if not masks:
            return f"错误: 未找到指定土层 {ids}。"
        row_mask = np.any(masks, axis=0)
    else:
        row_mask = np.ones(_hs_data.height, dtype=bool)

    subset = _hs_data.data[:, row_mask, :].copy()
    bands, h, w = subset.shape
    spectra = subset.reshape(bands, -1).T  # [N_pixels, Bands]

    smoothed = savgol_filter(spectra, window_length, polyorder, axis=1)

    smoothed_3d = smoothed.T.reshape(bands, h, w)
    _hs_data.data[:, row_mask, :] = smoothed_3d

    layer_desc = f"土层 {layer_ids}" if layer_ids else "所有土层"
    return (
        f"SG 平滑完成 ({layer_desc})。\n"
        f"  窗口: {window_length}, 多项式阶数: {polyorder}\n"
        f"  处理像元数: {spectra.shape[0]}\n"
        f"  平滑后均值: {smoothed.mean():.4f}, 标准差: {smoothed.std():.4f}"
    )


# ═══════════════════════════════════════════════════════════════
# 工具 9: 空间丰度热力图
# ═══════════════════════════════════════════════════════════════

def render_heatmap(output_path: str = "", band_index: int = 0, layer_id: str = "") -> str:
    """
    绘制指定波段/图层的空间丰度热力图。

    参数:
        output_path : 输出图片路径 (留空自动生成)
        band_index  : 使用的光谱波段索引 (0-based, 默认第1波段)
        layer_id    : 限制到指定土层 (留空则整幅图)

    返回:
        输出文件路径
    """
    if _hs_data.data is None:
        return "错误: 尚未加载 TIFF 数据。"

    if not output_path:
        os.makedirs("output", exist_ok=True)
        output_path = f"output/heatmap_band{band_index + 1}.png"

    # 选择数据范围
    if layer_id:
        for ly in _hs_data.layers:
            if str(ly.get("layer_id", "")) == str(layer_id):
                top, bottom = _hs_data.get_layer_pixel_range(ly)
                img = _hs_data.data[band_index, top:bottom, :]
                title = f"Band {band_index + 1} - Layer {layer_id} ({ly['name']})"
                break
        else:
            return f"错误: 未找到土层 '{layer_id}'。"
    else:
        img = _hs_data.data[band_index, :, :]
        title = f"Band {band_index + 1} - Full Profile"

    fig, ax = plt.subplots(figsize=(8, 6))
    im = ax.imshow(img, cmap="viridis", aspect="auto", origin="upper")
    ax.set_title(title, fontsize=12)
    ax.set_xlabel("Width (pixels)")
    ax.set_ylabel("Depth (pixels)")
    plt.colorbar(im, ax=ax, label="Reflectance")
    plt.tight_layout()
    fig.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close(fig)

    return f"热力图已保存: {output_path}"


# ═══════════════════════════════════════════════════════════════
# 工具 10: 光谱响应曲线
# ═══════════════════════════════════════════════════════════════

def render_spectrum(output_path: str = "", row: int = -1, col: int = -1,
                    layer_id: str = "", compare: bool = False) -> str:
    """
    绘制像元或土层的平均光谱响应曲线。

    参数:
        output_path : 输出图片路径
        row, col    : 像元坐标（-1 表示使用土层均值）
        layer_id    : 土层编号
        compare     : 是否绘制所有土层的对比曲线
    """
    if _hs_data.data is None:
        return "错误: 尚未加载 TIFF 数据。"

    if not output_path:
        os.makedirs("output", exist_ok=True)
        output_path = "output/spectrum.png"

    fig, ax = plt.subplots(figsize=(10, 5))
    bands_axis = np.arange(1, _hs_data.band_count + 1)

    if compare:
        # 绘制所有土层的平均光谱
        colors = plt.cm.tab10(np.linspace(0, 1, max(_hs_data.layer_count(), 1)))
        for i, ly in enumerate(_hs_data.layers):
            top, bottom = _hs_data.get_layer_pixel_range(ly)
            mean_spec = _hs_data.data[:, top:bottom, :].mean(axis=(1, 2))
            ax.plot(bands_axis, mean_spec, color=colors[i % 10],
                    label=f"L{ly['layer_id']}: {ly['name']}")
        ax.legend(fontsize=8, loc="upper right")
        title = "Layer Mean Spectral Comparison"
    elif layer_id:
        for ly in _hs_data.layers:
            if str(ly.get("layer_id", "")) == str(layer_id):
                top, bottom = _hs_data.get_layer_pixel_range(ly)
                mean_spec = _hs_data.data[:, top:bottom, :].mean(axis=(1, 2))
                ax.plot(bands_axis, mean_spec, "b-", linewidth=1.5)
                title = f"Layer {layer_id} ({ly['name']}) Mean Spectrum"
                break
        else:
            return f"错误: 未找到土层 '{layer_id}'。"
    elif row >= 0 and col >= 0:
        spectrum = _hs_data.data[:, row, col]
        ax.plot(bands_axis, spectrum, "r-", linewidth=1)
        title = f"Pixel ({row}, {col}) Spectrum"
    else:
        # 全图均值
        mean_spec = _hs_data.data.mean(axis=(1, 2))
        ax.plot(bands_axis, mean_spec, "g-", linewidth=1.5)
        title = "Full Profile Mean Spectrum"

    ax.set_title(title, fontsize=12)
    ax.set_xlabel("Band Index")
    ax.set_ylabel("Reflectance")
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    fig.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close(fig)

    return f"光谱图已保存: {output_path}"


# ═══════════════════════════════════════════════════════════════
# 工具 11: 分层统计
# ═══════════════════════════════════════════════════════════════

def compute_layer_stats(layer_id: str = "") -> str:
    """
    计算指定土层（或全部）的统计特征。

    返回:
        每个土层的波段均值、标准差、变异系数等
    """
    if _hs_data.data is None:
        return "错误: 尚未加载 TIFF 数据。"

    if layer_id:
        targets = [ly for ly in _hs_data.layers if str(ly.get("layer_id", "")) == str(layer_id)]
        if not targets:
            return f"错误: 未找到土层 '{layer_id}'。"
    else:
        targets = _hs_data.layers

    lines = [f"剖面 {_hs_data.profile_id} 分层统计报告:", "=" * 50]
    lines.append(f"{'土层':<6} {'名称':<12} {'深度(cm)':<12} {'均值':<10} {'标准差':<10} {'变异系数':<10}")

    for ly in targets:
        top, bottom = _hs_data.get_layer_pixel_range(ly)
        layer_data = _hs_data.data[:, top:bottom, :]
        mean_val = layer_data.mean()
        std_val = layer_data.std()
        cv = std_val / mean_val if mean_val > 0 else 0

        depth_range = f"{ly['top_cm']}-{ly['bottom_cm']}"
        lines.append(
            f"L{ly['layer_id']:<5} {ly['name']:<12} {depth_range:<12} "
            f"{mean_val:<10.4f} {std_val:<10.4f} {cv:<10.4f}"
        )

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════
# 工具注册表
# ═══════════════════════════════════════════════════════════════

TOOL_REGISTRY = {
    "load_tiff": load_tiff,
    "load_from_folder": load_from_folder,
    "set_meta_file": set_meta_file,
    "get_layer_info": get_layer_info,
    "extract_layer": extract_layer,
    "extract_spectrum": extract_spectrum,
    "clean_layer_noise": clean_layer_noise,
    "msc_transform": msc_transform,
    "snv_transform": snv_transform,
    "sg_smooth": sg_smooth,
    "render_heatmap": render_heatmap,
    "render_spectrum": render_spectrum,
    "compute_layer_stats": compute_layer_stats,
}

# 工具描述（供 LLM 决策使用）
TOOL_DESCRIPTIONS = {
    "load_tiff": "加载高光谱 TIFF 文件和分层元数据。参数: tif_path (TIFF路径), meta_path (分层元数据路径, .xlsx/.json/.txt)",
    "load_from_folder": "选择剖面文件夹，自动用文件夹名匹配编号并寻找TIF。参数: folder_path (文件夹路径，文件夹名即为剖面编号)",
    "set_meta_file": "设置全局分层元数据文件路径。参数: meta_path (分层元数据文件路径)",
    "get_layer_info": "获取当前加载数据的所有分层信息摘要。无需参数。",
    "extract_layer": "提取指定土层的像素数据并返回统计。参数: layer_id (土层编号, 如'1')",
    "extract_spectrum": "提取指定像元的光谱曲线。参数: row (行号), col (列号)",
    "clean_layer_noise": "对指定土层进行噪声清洗。参数: layer_id (土层编号), strategy (策略: median/mean/none)",
    "msc_transform": "多元散射校正 MSC。参数: layer_ids (逗号分隔的土层编号, 如'1,2', 留空=全部)",
    "snv_transform": "标准正态变量变换 SNV。参数: layer_ids (逗号分隔的土层编号, 留空=全部)",
    "sg_smooth": "Savitzky-Golay 光谱平滑。参数: window_length (窗口大小, 默认7), polyorder (多项式阶数, 默认2), layer_ids",
    "render_heatmap": "绘制空间丰度热力图。参数: output_path, band_index (波段序号, 0-based), layer_id",
    "render_spectrum": "绘制光谱响应曲线。参数: output_path, row, col, layer_id, compare (是否对比所有层)",
    "compute_layer_stats": "计算分层统计特征（均值、标准差、变异系数）。参数: layer_id (留空=全部)",
}
